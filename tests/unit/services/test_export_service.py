"""ExportService の単体テスト。"""

import io

import pytest
from openpyxl import load_workbook

from src.services.export_service import ExportService


@pytest.mark.parametrize(
    ('headers', 'rows', 'expected_csv'),
    [
        (
            ['header1', 'header2'],
            [['val1', 'val2'], ['val3', 'val4']],
            'header1,header2\r\nval1,val2\r\nval3,val4\r\n',
        ),
        (['header1'], [['val1']], 'header1\r\nval1\r\n'),
        ([], [], '\r\n'),
        (['header1', 'header2'], [], 'header1,header2\r\n'),
    ],
)
def test_create_csvが正しくCSVを作成する(
    headers: list[str], rows: list[list[str]], expected_csv: str
) -> None:
    # Act
    result = ExportService.create_csv(headers, rows)

    # Assert
    assert isinstance(result, io.StringIO)
    result.seek(0)
    assert result.read() == expected_csv


def test_create_excelが正しくExcelファイルを作成する() -> None:
    # Arrange
    headers = ['header1', 'header2']
    rows = [['val1', 'val2'], ['val3', 'val4']]

    # Act
    result = ExportService.create_excel(headers, rows)

    # Assert
    assert isinstance(result, io.BytesIO)
    assert result.tell() == 0  # seek(0) が呼ばれていることを確認

    # Excelファイルの内容を確認
    result.seek(0)
    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None
    assert ws.title == '生成結果'

    # ヘッダー行を確認
    header_row = [cell.value for cell in ws[1]]
    assert header_row == headers

    # データ行を確認
    data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    assert data_rows == rows


def test_create_excelが空のデータでも動作する() -> None:
    # Arrange
    headers = ['header1', 'header2']
    rows: list[list[str]] = []

    # Act
    result = ExportService.create_excel(headers, rows)

    # Assert
    assert isinstance(result, io.BytesIO)

    result.seek(0)
    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None

    # ヘッダー行のみ存在することを確認
    header_row = [cell.value for cell in ws[1]]
    assert header_row == headers
    assert ws.max_row == 1
